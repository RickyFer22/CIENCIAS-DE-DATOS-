import sqlite3
from openpyxl import Workbook
import openpyxl.utils # For get_column_letter
import os
import logging
import sys

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(filename)s - [%(lineno)d] - %(message)s')

def sanitize_sheet_name(name, existing_names, counter=0):
    """Sanitizes a table name to be a valid Excel sheet name, ensuring uniqueness."""
    # Characters not allowed in Excel sheet names by openpyxl or common sense
    invalid_chars = ['[', ']', '*', ':', '?', '/', '\\']
    for char in invalid_chars:
        name = name.replace(char, '_') # Replace with underscore or remove
    
    # Truncate to 30 characters (Excel limit is 31, but be safe with potential suffixes)
    name = name[:30]

    if not name: # If name becomes empty after sanitization
        name = f"Table_{counter}"
        
    # Ensure uniqueness
    original_name = name
    suffix = 1
    while name in existing_names:
        name = f"{original_name[:28]}_{suffix}" # Truncate original to make space for suffix
        suffix += 1
    return name

def main():
    logging.info("Script started: 6listas_en_excel.py")
    
    current_folder = os.path.dirname(os.path.abspath(__file__))
    logging.info(f"Script directory: {current_folder}")

    db_filename = 'precios_competencia.db'
    db_path = os.path.join(current_folder, db_filename)
    excel_filename = 'listas_de_precios.xlsx'
    excel_file_path = os.path.join(current_folder, excel_filename)

    logging.info(f"Database path: {db_path}")
    logging.info(f"Excel output path: {excel_file_path}")

    if not os.path.isfile(db_path):
        logging.error(f"Database not found at: {db_path}. Please ensure the database file exists.")
        sys.exit(1)

    conn = None
    wb = Workbook() # Create workbook at the beginning
    # Remove the default sheet created by openpyxl, if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])
        logging.debug("Removed default 'Sheet' from new workbook.")

    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        logging.info(f"Successfully connected to database: {db_path}")

        c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
        tablas = c.fetchall()

        if not tablas:
            logging.info("No tables found in the database. An empty Excel file will be created.")
        
        processed_sheet_names = set()

        for i, tabla_row in enumerate(tablas):
            nombre_tabla_original = tabla_row[0]
            
            # Sanitize sheet name
            safe_sheet_name = sanitize_sheet_name(nombre_tabla_original, processed_sheet_names, i)
            processed_sheet_names.add(safe_sheet_name)
            
            logging.info(f"Processing table: '{nombre_tabla_original}' into sheet: '{safe_sheet_name}'")
            ws = wb.create_sheet(title=safe_sheet_name)

            try:
                # Use brackets for table names to handle special characters or keywords
                c.execute(f"SELECT * FROM [{nombre_tabla_original}]")
                rows = c.fetchall()
                
                if c.description: # Check if query returned columns
                    column_names = [description[0] for description in c.description]
                    ws.append(column_names) # Write headers
                else: # Should not happen with SELECT * if table exists and is not empty schema-wise
                    logging.warning(f"No column descriptions returned for table {nombre_tabla_original}. Sheet will be empty or only contain data if any.")
                    column_names = []


                if rows:
                    logging.info(f"Exporting {len(rows)} rows from table '{nombre_tabla_original}' to sheet '{safe_sheet_name}'.")
                    for row_data in rows:
                        ws.append(row_data)
                    
                    # Adjust column widths
                    for col_idx, column_letter in enumerate([openpyxl.utils.get_column_letter(j+1) for j in range(ws.max_column)], 1):
                        max_cell_length = 0
                        # Check header length
                        if column_names and col_idx <= len(column_names):
                             max_cell_length = len(str(column_names[col_idx-1]))
                        
                        for cell in ws[column_letter]: # Iterate through cells in the current column
                            try:
                                if cell.value is not None:
                                    cell_content_length = len(str(cell.value))
                                    if cell_content_length > max_cell_length:
                                        max_cell_length = cell_content_length
                            except Exception as e_cell_len:
                                logging.warning(f"Could not determine length for cell {cell.coordinate} in sheet '{safe_sheet_name}': {e_cell_len}")
                        
                        adjusted_width = (max_cell_length + 2) * 1.1 # Slightly smaller factor
                        if adjusted_width > 60: adjusted_width = 60 # Cap width
                        if adjusted_width < 8: adjusted_width = 8   # Min width
                        ws.column_dimensions[column_letter].width = adjusted_width
                    logging.debug(f"Adjusted column widths for sheet: {safe_sheet_name}")

                else:
                    logging.info(f"No data found in table '{nombre_tabla_original}'.")
                    if not column_names: # If there were no columns either (empty table schema)
                         ws.cell(row=1, column=1, value="Table is empty or has no data.")


            except sqlite3.Error as e_table_fetch:
                logging.error(f"Error fetching data from table '{nombre_tabla_original}': {e_table_fetch}")
                # Write error to sheet if possible
                try:
                    ws.cell(row=1, column=1, value=f"Error fetching data: {e_table_fetch}")
                except Exception as e_ws_err:
                    logging.error(f"Could not write error to sheet {safe_sheet_name}: {e_ws_err}")
                continue # Skip to next table

        # Ensure at least one sheet exists, otherwise save will fail for some libraries or be non-standard
        if not wb.sheetnames:
            wb.create_sheet(title="DefaultSheet")
            logging.info("No tables processed or all resulted in errors before sheet creation. Created a default sheet.")

        try:
            wb.save(excel_file_path)
            logging.info(f"Excel file saved successfully: {excel_file_path}")
        except Exception as e_excel_save:
            logging.error(f"Error saving Excel file {excel_file_path}: {e_excel_save}", exc_info=True)
            # Consider sys.exit(1) if saving is critical

    except sqlite3.Error as e_db:
        logging.error(f"Database connection or operational error: {e_db}", exc_info=True)
        sys.exit(1) # Exit if DB connection fails as it's fundamental
    except Exception as e_main:
        logging.error(f"An unexpected error occurred in main execution: {e_main}", exc_info=True)
        sys.exit(1)
    finally:
        if conn:
            try:
                conn.close()
                logging.info("Database connection closed.")
            except sqlite3.Error as e_db_close:
                logging.error(f"Error closing database connection: {e_db_close}", exc_info=True)
    
    logging.info("Script finished: 6listas_en_excel.py")

if __name__ == "__main__":
    main()
```
